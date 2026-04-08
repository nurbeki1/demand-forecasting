"""
Report Service for Demand Forecasting
Generates PDF and Excel reports with forecasts, analytics, and insights
"""

import os
import io
import logging
from datetime import datetime, timedelta
from typing import Optional, List, Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)


class ReportService:
    """Service for generating PDF and Excel reports"""

    def __init__(self):
        self.company_name = "Forecast AI"
        self.report_styles = {
            'header_fill': PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid"),
            'header_font': Font(color="FFFFFF", bold=True, size=12),
            'title_font': Font(bold=True, size=14),
            'subtitle_font': Font(bold=True, size=11),
            'normal_font': Font(size=10),
            'number_font': Font(size=10),
            'positive_font': Font(color="10B981", bold=True),
            'negative_font': Font(color="EF4444", bold=True),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            ),
        }

    async def generate_daily_report_excel(self, data: Dict[str, Any]) -> bytes:
        """Generate daily Excel report"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Daily Report"

        # Title
        ws['A1'] = f"{self.company_name} - Daily Report"
        ws['A1'].font = self.report_styles['title_font']
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A2'].font = self.report_styles['normal_font']

        # Summary Section
        ws['A4'] = "📊 Summary"
        ws['A4'].font = self.report_styles['subtitle_font']

        summary = data.get('summary', {})
        row = 5
        for key, value in summary.items():
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            row += 1

        # Top Products Section
        ws[f'A{row + 1}'] = "🏆 Top Products"
        ws[f'A{row + 1}'].font = self.report_styles['subtitle_font']

        products = data.get('top_products', [])
        if products:
            headers = ['Product', 'Sales', 'Revenue', 'Trend']
            row += 2
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = self.report_styles['header_fill']
                cell.font = self.report_styles['header_font']
                cell.border = self.report_styles['border']

            for product in products[:10]:
                row += 1
                ws.cell(row=row, column=1, value=product.get('name', ''))
                ws.cell(row=row, column=2, value=product.get('sales', 0))
                ws.cell(row=row, column=3, value=product.get('revenue', 0))
                ws.cell(row=row, column=4, value=product.get('trend', ''))
                for col in range(1, 5):
                    ws.cell(row=row, column=col).border = self.report_styles['border']

        # Alerts Section
        ws[f'A{row + 2}'] = "⚠️ Active Alerts"
        ws[f'A{row + 2}'].font = self.report_styles['subtitle_font']

        alerts = data.get('alerts', [])
        row += 3
        for alert in alerts[:5]:
            ws[f'A{row}'] = f"• {alert.get('message', '')}"
            ws[f'A{row}'].font = self.report_styles['normal_font']
            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    async def generate_forecast_report_excel(
        self,
        product_id: str,
        history: List[Dict],
        forecast: List[Dict],
        insights: Dict[str, Any]
    ) -> bytes:
        """Generate forecast Excel report for a specific product"""
        wb = Workbook()

        # Sheet 1: Overview
        ws1 = wb.active
        ws1.title = "Overview"

        ws1['A1'] = f"{self.company_name} - Forecast Report"
        ws1['A1'].font = self.report_styles['title_font']
        ws1['A2'] = f"Product: {product_id}"
        ws1['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

        # Summary metrics
        ws1['A5'] = "📈 Forecast Summary"
        ws1['A5'].font = self.report_styles['subtitle_font']

        if forecast:
            avg_forecast = sum(f.get('predicted_demand', 0) for f in forecast) / len(forecast)
            max_forecast = max(f.get('predicted_demand', 0) for f in forecast)
            min_forecast = min(f.get('predicted_demand', 0) for f in forecast)

            ws1['A6'] = "Average Predicted Demand:"
            ws1['B6'] = round(avg_forecast, 2)
            ws1['A7'] = "Max Predicted Demand:"
            ws1['B7'] = round(max_forecast, 2)
            ws1['A8'] = "Min Predicted Demand:"
            ws1['B8'] = round(min_forecast, 2)
            ws1['A9'] = "Forecast Period:"
            ws1['B9'] = f"{len(forecast)} days"

        # Insights
        if insights:
            ws1['A11'] = "💡 Key Insights"
            ws1['A11'].font = self.report_styles['subtitle_font']

            row = 12
            if insights.get('summary'):
                ws1[f'A{row}'] = insights['summary']
                row += 2

            if insights.get('risk_level'):
                ws1[f'A{row}'] = f"Risk Level: {insights['risk_level']}"
                row += 1

            actions = insights.get('action_items', [])
            if actions:
                ws1[f'A{row}'] = "Recommended Actions:"
                row += 1
                for action in actions[:5]:
                    ws1[f'A{row}'] = f"• {action.get('action', '')}"
                    row += 1

        # Sheet 2: Historical Data
        ws2 = wb.create_sheet("Historical Data")
        ws2['A1'] = "Historical Demand Data"
        ws2['A1'].font = self.report_styles['subtitle_font']

        headers = ['Date', 'Demand', 'Inventory', 'Price']
        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=2, column=col, value=header)
            cell.fill = self.report_styles['header_fill']
            cell.font = self.report_styles['header_font']
            cell.border = self.report_styles['border']

        for i, record in enumerate(history[-30:], 3):  # Last 30 days
            ws2.cell(row=i, column=1, value=record.get('date', ''))
            ws2.cell(row=i, column=2, value=record.get('demand', 0))
            ws2.cell(row=i, column=3, value=record.get('inventory', 0))
            ws2.cell(row=i, column=4, value=record.get('price', 0))
            for col in range(1, 5):
                ws2.cell(row=i, column=col).border = self.report_styles['border']

        # Sheet 3: Forecast Data
        ws3 = wb.create_sheet("Forecast")
        ws3['A1'] = "Demand Forecast"
        ws3['A1'].font = self.report_styles['subtitle_font']

        headers = ['Date', 'Predicted Demand', 'Lower Bound', 'Upper Bound']
        for col, header in enumerate(headers, 1):
            cell = ws3.cell(row=2, column=col, value=header)
            cell.fill = self.report_styles['header_fill']
            cell.font = self.report_styles['header_font']
            cell.border = self.report_styles['border']

        for i, record in enumerate(forecast, 3):
            ws3.cell(row=i, column=1, value=record.get('date', ''))
            ws3.cell(row=i, column=2, value=record.get('predicted_demand', 0))
            ws3.cell(row=i, column=3, value=record.get('lower_bound', 0))
            ws3.cell(row=i, column=4, value=record.get('upper_bound', 0))
            for col in range(1, 5):
                ws3.cell(row=i, column=col).border = self.report_styles['border']

        # Add chart
        if len(forecast) > 1:
            chart = LineChart()
            chart.title = "Demand Forecast"
            chart.x_axis.title = "Date"
            chart.y_axis.title = "Demand"

            data = Reference(ws3, min_col=2, min_row=2, max_row=2 + len(forecast))
            dates = Reference(ws3, min_col=1, min_row=3, max_row=2 + len(forecast))
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(dates)
            chart.width = 15
            chart.height = 8

            ws3.add_chart(chart, "F2")

        # Adjust column widths
        for ws in [ws1, ws2, ws3]:
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    async def generate_analytics_report_excel(self, analytics_data: Dict[str, Any]) -> bytes:
        """Generate analytics Excel report"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Analytics Report"

        # Title
        ws['A1'] = f"{self.company_name} - Analytics Report"
        ws['A1'].font = self.report_styles['title_font']
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

        # Dataset Summary
        ws['A4'] = "📊 Dataset Summary"
        ws['A4'].font = self.report_styles['subtitle_font']

        summary = analytics_data.get('summary', {})
        row = 5
        summary_items = [
            ('Total Products', summary.get('total_products', 0)),
            ('Total Records', summary.get('total_records', 0)),
            ('Date Range', f"{summary.get('date_min', '')} to {summary.get('date_max', '')}"),
            ('Average Demand', round(summary.get('avg_demand', 0), 2)),
            ('Total Revenue', f"${summary.get('total_revenue', 0):,.2f}"),
        ]

        for label, value in summary_items:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1

        # Trending Products
        ws[f'A{row + 1}'] = "📈 Trending Up"
        ws[f'A{row + 1}'].font = self.report_styles['subtitle_font']

        trending_up = analytics_data.get('trending_up', [])
        row += 2
        for product in trending_up[:5]:
            ws[f'A{row}'] = product.get('product_id', '')
            ws[f'B{row}'] = f"+{product.get('growth', 0):.1f}%"
            ws[f'B{row}'].font = self.report_styles['positive_font']
            row += 1

        # Declining Products
        ws[f'A{row + 1}'] = "📉 Trending Down"
        ws[f'A{row + 1}'].font = self.report_styles['subtitle_font']

        trending_down = analytics_data.get('trending_down', [])
        row += 2
        for product in trending_down[:5]:
            ws[f'A{row}'] = product.get('product_id', '')
            ws[f'B{row}'] = f"{product.get('growth', 0):.1f}%"
            ws[f'B{row}'].font = self.report_styles['negative_font']
            row += 1

        # Category Performance
        ws[f'A{row + 1}'] = "📦 Category Performance"
        ws[f'A{row + 1}'].font = self.report_styles['subtitle_font']

        categories = analytics_data.get('categories', [])
        row += 2
        headers = ['Category', 'Products', 'Avg Demand', 'Revenue']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = self.report_styles['header_fill']
            cell.font = self.report_styles['header_font']

        for cat in categories[:10]:
            row += 1
            ws.cell(row=row, column=1, value=cat.get('category', ''))
            ws.cell(row=row, column=2, value=cat.get('product_count', 0))
            ws.cell(row=row, column=3, value=round(cat.get('avg_demand', 0), 2))
            ws.cell(row=row, column=4, value=f"${cat.get('revenue', 0):,.2f}")

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    async def generate_kz_market_report_excel(self, kz_data: Dict[str, Any]) -> bytes:
        """Generate Kazakhstan market analysis Excel report"""
        wb = Workbook()
        ws = wb.active
        ws.title = "KZ Market Analysis"

        # Title
        ws['A1'] = f"{self.company_name} - Kazakhstan Market Report"
        ws['A1'].font = self.report_styles['title_font']
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

        # Product Info
        product = kz_data.get('product', {})
        ws['A4'] = "📦 Product Analysis"
        ws['A4'].font = self.report_styles['subtitle_font']
        ws['A5'] = f"Product: {product.get('name', 'N/A')}"
        ws['A6'] = f"Wholesale Price: ${product.get('wholesale_price', 0):.2f}"
        ws['A7'] = f"Category: {product.get('category', 'N/A')}"

        # City Profitability
        ws['A9'] = "🏙️ City Profitability Analysis"
        ws['A9'].font = self.report_styles['subtitle_font']

        cities = kz_data.get('cities', [])
        headers = ['City', 'Tier', 'Retail Price', 'Profit', 'Margin %', 'Recommendation']
        row = 10
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = self.report_styles['header_fill']
            cell.font = self.report_styles['header_font']
            cell.border = self.report_styles['border']

        for city in cities:
            row += 1
            ws.cell(row=row, column=1, value=city.get('city', ''))
            ws.cell(row=row, column=2, value=city.get('tier', ''))
            ws.cell(row=row, column=3, value=f"${city.get('retail_price', 0):.2f}")
            ws.cell(row=row, column=4, value=f"${city.get('profit', 0):.2f}")

            margin = city.get('margin', 0)
            margin_cell = ws.cell(row=row, column=5, value=f"{margin:.1f}%")
            if margin >= 25:
                margin_cell.font = self.report_styles['positive_font']
            elif margin < 15:
                margin_cell.font = self.report_styles['negative_font']

            ws.cell(row=row, column=6, value=city.get('recommendation', ''))

            for col in range(1, 7):
                ws.cell(row=row, column=col).border = self.report_styles['border']

        # Summary
        ws[f'A{row + 2}'] = "📊 Summary"
        ws[f'A{row + 2}'].font = self.report_styles['subtitle_font']

        summary = kz_data.get('summary', {})
        ws[f'A{row + 3}'] = f"Best City: {summary.get('best_city', 'N/A')}"
        ws[f'A{row + 4}'] = f"Average Margin: {summary.get('avg_margin', 0):.1f}%"
        ws[f'A{row + 5}'] = f"Recommended Cities: {summary.get('recommended_count', 0)}"

        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 25

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()


# Global instance
report_service = ReportService()


def get_report_service() -> ReportService:
    """Get report service instance"""
    return report_service